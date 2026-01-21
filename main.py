import struct
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# --- GUI ---
import tkinter as tk
from tkinter import filedialog, messagebox

PID_IDS = {
    0x27: "PV",
    0x28: "SP",
    0x29: "CV",
    0x30: "CV_P",
    0x31: "CV_I",
    0x32: "CV_D",
    0x33: "Kp",
    0x34: "Ki",
    0x35: "Kd",
}

PLOT_SIGNALS = ["CV", "CV_P", "CV_I", "CV_D", "PV", "SP"]


def load_table(path: Path) -> pd.DataFrame:
    suf = path.suffix.lower()
    if suf == ".csv":
        return pd.read_csv(path)
    if suf in (".xls", ".xlsx"):
        return pd.read_excel(path)
    raise ValueError("Поддерживаются только CSV и Excel")


def bytes_list_to_float_le(hex_bytes_4) -> float:
    b = bytes(int(x, 16) for x in hex_bytes_4)
    return struct.unpack("<f", b)[0]


def add_one_chart(ws, title, x_col_idx, y_col_indices, anchor_cell):
    max_row = ws.max_row
    if max_row < 3 or not y_col_indices:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "t, s"

    xvalues = Reference(ws, min_col=x_col_idx, min_row=2, max_row=max_row)

    for col in y_col_indices:
        values = Reference(ws, min_col=col, min_row=1, max_row=max_row)  # с заголовком
        chart.add_data(values, titles_from_data=True)
        chart.series[-1].xvalues = xvalues

    ws.add_chart(chart, anchor_cell)


def convert_file(input_file: Path, output_file: Path):
    df = load_table(input_file)

    if "Data" not in df.columns:
        raise ValueError("Не найдена колонка 'Data' во входном файле.")
    if "Timestamp" not in df.columns:
        raise ValueError("Не найдена колонка 'Timestamp' во входном файле.")

    ts = pd.to_datetime(df["Timestamp"], errors="coerce")

    rows = []
    for data_str, t in zip(df["Data"].astype(str), ts):
        if pd.isna(t):
            continue
        parts = data_str.strip().split()
        if len(parts) < 5:
            continue

        msg_id = int(parts[0], 16)
        if msg_id not in PID_IDS:
            continue

        value = bytes_list_to_float_le(parts[-4:])
        rows.append({"Timestamp": t, "Signal": PID_IDS[msg_id], "Value": value})

    decoded = pd.DataFrame(rows).sort_values("Timestamp")
    if decoded.empty:
        raise ValueError("Не удалось декодировать ни одной записи. Проверь формат 'Data' и ID сообщений.")

    # t, сек от начала
    t0 = decoded["Timestamp"].iloc[0]
    decoded["TimeSec"] = (decoded["Timestamp"] - t0).dt.total_seconds()

    # wide-таблица по времени + заполнение последним значением
    wide = decoded.pivot_table(
        index="TimeSec", columns="Signal", values="Value", aggfunc="last"
    ).sort_index().ffill()

    out_df = wide.reset_index()

    # Запись в Excel
    sheet_name = "data"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Добавляем один график
    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    x_col = headers["TimeSec"]
    y_cols = [headers[s] for s in PLOT_SIGNALS if s in headers]

    add_one_chart(
        ws,
        "PID signals: CV, CV_P, CV_I, CV_D, PV, SP",
        x_col,
        y_cols,
        "L2"
    )

    wb.save(output_file)


def main():
    root = tk.Tk()
    root.withdraw()  # не показываем пустое главное окно

    in_path = filedialog.askopenfilename(
        title="Выберите входной файл (CSV или Excel)",
        filetypes=[
            ("CSV files", "*.csv"),
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*"),
        ],
    )
    if not in_path:
        return

    out_path = filedialog.asksaveasfilename(
        title="Сохранить результат как...",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="pid_decoded_with_charts.xlsx",
    )
    if not out_path:
        return

    try:
        convert_file(Path(in_path), Path(out_path))
    except Exception as e:
        messagebox.showerror("Ошибка", str(e))
        return

    messagebox.showinfo("Готово", f"Файл создан:\n{out_path}")


if __name__ == "__main__":
    main()
